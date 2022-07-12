from email import header
from operator import imod
import requests
from urllib import request
from urllib import parse
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import numpy as np

url = 'https://www.lagou.com/jobs/v2/positionAjax.json'

def writeExcel(json):
    wb = Workbook()
    wb.title = 'Sheet1'
    wb.create_sheet('sheet1',index=1)

    sheet = wb.get_sheet_by_name('Sheet')
    # row = [1 ,2, 3, 4, 5]
    # sheet.append(row)
    label = [[0],
     [1],
     [2],
     [3]
     ]
    feature = [
        [0.1, 0.2, 0.3, 0.4, 0.5],
        [0.11, 0.21, 0.31, 0.41, 0.51],
        [0.6, 0.7, 0.8, 0.9, 1.00],
        [1.1, 1.2, 1.3, 1.4, 1.5],
        ]
    label = np.array(label)
    feature = np.array(feature)
 
    label_input = []
    for l in range(len(label)):
        label_input.append(label[l][0])
    sheet.append(label_input)
    for f in range(len(feature[0])):
        sheet.append(feature[:, f].tolist())
    sheet['A2'] = '11'
    wb.save(r'D:\导出.xlsx')
    print(wb.get_sheet_names())

def getExcelJson(path):
    sessionId = ''
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            # 获得所有sheet的名称
            print(wb.get_sheet_names()) 
            #按工作簿定位工作表
            a_sheet = wb.get_sheet_by_name('Sheet1') 
            # 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
            sheet = wb.active
            # 获取某个单元格的值
            for i in range(sheet.max_row):
                celA = sheet['A'+ str(i+1)]
                celB = sheet['B' + str(i+1)]
                if celB.value != None:
                    keyv = celA.value + '=' + str(celB.value)
                    sessionId += ';' + keyv
            return sessionId[1:]
        except:
            print('发生异常')
            return sessionId
    else:
        print("文件不存在")
        return ''

def get_json(url, num,sessionId):
    """
    从指定的url中通过requests请求携带请求头和请求体获取网页中的信息,
    :return:
    """
    headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    'Host': 'www.lagou.com',
    'Origin': 'https://www.lagou.com',
    'Referer': 'https://www.lagou.com/wn/jobs?labelWords=&fromSearch=true&suginput=&kd=%25E4%25BA%25A7%25E5%2593%2581%25E7%25BB%258F%25E7%2590%2586',
    'X-Anit-Forge-Code': '0',
    'X-Anit-Forge-Token': 'None',
    'X-Requested-With': 'XMLHttpRequest',
    'Cookie': sessionId
    }
    
    data = {
    'first': 'true',
    'pn': num,
    'kd': 'python工程师'}
    requ = request.Request(url,headers=headers,data=parse.urlencode(data).encode('utf-8'),method='POST')
    result = request.urlopen(requ)
    print('请求结果：',result.read().decode('utf-8'))


 
def postGetData(url,headers):
    s = requests.Session()
    print('建立session：', s, '\n\n')
    s.get(url=url, headers=headers, timeout=3)
    cookie = s.cookies
    print('获取cookie：', cookie, '\n\n')
    res = requests.post(url, headers=headers, data=data, cookies=cookie, timeout=3)
    print('返回状态:',res.status_code)
    res.raise_for_status()
    res.encoding = 'utf-8'
    page_data = res.json()
    print('请求响应结果：', page_data, '\n\n')
    return page_data
 
sessionId = getExcelJson('D:\\file\\数据1.xlsx')
writeExcel('f')
if sessionId.strip() == '':
    print("未获取到sessionId")
else:
 print(get_json(url, 1,sessionId))
